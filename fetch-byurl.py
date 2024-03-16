from openpyxl import load_workbook
import requests
import json
import sys

wb = load_workbook('hostname-sheet-retrieve/hostnames.xlsx')
ws = wb.active
urls = []

for i in range(1, 201):
    cell = ws.cell(row = i, column = 1).value
    if cell:
        urls.append(cell)

wb.close()

for j, url in enumerate(urls):
    file_path = 'hostname-sheet-retrieve/output/final{}.json'.format(j)
    with open(file_path, 'w') as outfile:
        response = requests.get(url)
        if response.status_code == 200:
            json.dump(response.json(), outfile, indent = 4)
    
sys.exit()