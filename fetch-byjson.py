from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import load_workbook
import requests
import json
import sys
import os


def retrieve_hostnames_from_excel():
    wb = load_workbook('hostname-sheet-retrieve/inputhostnames.xlsx')
    ws = wb.active
    hostnames = []

    for i in range(1, 201):
        cell = ws.cell(row=i, column=1).value
        if cell:
            hostnames.append(cell)

    wb.close()
    return hostnames


def send_post_requests(hostnames):
    url = 'https://jsonplaceholder.typicode.com/posts'

    for j, hostname in enumerate(hostnames):
        with open('hostname-sheet-retrieve/body.json') as x:
            json_data = json.load(x)
        
        for data in json_data:
            data['id'] = hostname
            
        response = requests.post(url, json=json_data)
        file_path = f'hostname-sheet-retrieve/output-json/final{j}.json'

        with open(file_path, 'w') as outfile:
            if response.status_code == 201:
                json.dump(response.json(), outfile, indent=4)
            else:
                print("Request failed")


def convert_json_to_pdf(json_folder_path, output_pdf_folder):
    for file_name in os.listdir(json_folder_path):
        if file_name.endswith('.json'):
            input_json_path = os.path.join(json_folder_path, file_name)
            output_pdf_path = os.path.join(output_pdf_folder, os.path.splitext(file_name)[0] + '.pdf')
            
            with open(input_json_path, 'r') as json_file:
                data = json.load(json_file)
                pretty_json = json.dumps(data, indent=4)
                c = canvas.Canvas(output_pdf_path, pagesize=letter)
                
                c.setFillColorRGB(0, 0, 0)
                c.rect(0, 0, letter[0], letter[1], fill=1)
                c.setFillColorRGB(1, 1, 1)

                y_offset = 750

                for line in pretty_json.split('\n'):
                    c.drawString(10, y_offset, line)
                    y_offset -= 20
                    
                    if y_offset <= 50:
                        c.showPage()
                        c.setFont("Helvetica", 12)
                        y_offset = 750
                c.save()


def main():
    json_folder_path = 'hostname-sheet-retrieve/output-json'
    output_pdf_folder = 'hostname-sheet-retrieve/output-pdf'
    
    hostnames = retrieve_hostnames_from_excel()
    send_post_requests(hostnames)
    convert_json_to_pdf(json_folder_path, output_pdf_folder)


if __name__ == "__main__":
    main()
    sys.exit()
